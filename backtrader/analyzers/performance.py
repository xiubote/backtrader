# -*- coding: utf-8 -*-
"""
Created on Sun Jan 12 18:56:13 2020

@author: ZifengWang
"""

from __future__ import (absolute_import, division, print_function,
                        unicode_literals)

from collections import OrderedDict
from backtrader import Analyzer
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import os

class Performance(Analyzer):

    params = ( )

    def __init__(self):
        
        self.dataclose = self.datas[0].close
        self.pass_count = 0
        #self.frequency = self.datas[0].params.frequency
        self.frequency = 'min'
        try:
            os.mkdir('D:\\work')
        except FileExistsError:
            pass

    def start(self):
        super().start()
        
        self.tradetime = []
        #self.tradetimebar = []
        self.position_bar = {}
        self.position_daily = {}
        self.cash_bar = {}
        self.cash_daily = {}
        self.value_bar = OrderedDict()
        self.value_daily = OrderedDict()
        self.net_bar = {}
        self.net_daily = {}
        self.index_bar = {}
        self.index_daily = {}
        self.Tradetotal = 0
        self.position_cw = {}
        self.drawdownX = []
        self.trade_image = []
        self.order_image = {}
        self.order_image_cum = {}
        self.signal_image = {}
        self.trade_count = 0
        
    def notify_order(self, order):

        if order.status in [order.Submitted, order.Accepted]:
            if order.isbuy():
                self.signal_image[order.created.dt] = [order.ref,
                     order.created.size]
            else:
                self.signal_image[order.created.dt] = [order.ref,
                     order.created.size]

        if order.status in [order.Completed]:
            if self.frequency == 'day':
                if str(self.datas[0].datetime.date(0))+str(order.data._name) not in self.order_image.keys():
#                try:
#                    self.order_image[str(self.datas[0].datetime.date(0))+str(order.data._name)] = self.order_image[str(self.datas[0].datetime.date(0))+str(order.data._name)]
#                except KeyError:
                    self.order_image[str(self.datas[0].datetime.date(0))+str(order.data._name)] = [self.datas[0].datetime.date(0),
                         order.data._name,
                         order.executed.price,
                         order.executed.size,
                         order.executed.comm,
                         order.executed.value,
                         self.strategy.broker.getposition(order.data).size
                         ]
                    self.order_image_cum[str(self.datas[0].datetime.date(0))+' '+str(order.data._name)] = [self.datas[0].datetime.date(0), order.data._name, self.strategy.broker.getposition(order.data).size]                    
                else:
                    self.order_image[str(self.datas[0].datetime.date(0))+str(order.data._name)][3] += order.executed.size
                    self.order_image[str(self.datas[0].datetime.date(0))+str(order.data._name)][4] += order.executed.comm
                    self.order_image[str(self.datas[0].datetime.date(0))+str(order.data._name)][5] += order.executed.value
                    self.order_image[str(self.datas[0].datetime.date(0))+str(order.data._name)][6] = self.strategy.broker.getposition(order.data).size

            else:
                if str(self.datas[0].datetime.datetime(0))+str(order.data._name) not in self.order_image.keys():
                    self.order_image[str(self.datas[0].datetime.datetime(0))+str(order.data._name)] = [self.datas[0].datetime.datetime(0),
                         order.data._name,
                         order.executed.price,
                         order.executed.size,
                         order.executed.comm,
                         order.executed.value,
                         self.strategy.broker.getposition(order.data).size
                         ]
                    self.order_image_cum[str(self.datas[0].datetime.datetime(0))+' '+str(order.data._name)] = [self.datas[0].datetime.datetime(0), order.data._name, self.strategy.broker.getposition(order.data).size]
                else:
                    self.order_image[str(self.datas[0].datetime.datetime(0))+str(order.data._name)][3] += order.executed.size
                    self.order_image[str(self.datas[0].datetime.datetime(0))+str(order.data._name)][4] += order.executed.comm
                    self.order_image[str(self.datas[0].datetime.datetime(0))+str(order.data._name)][5] += order.executed.value
                    self.order_image[str(self.datas[0].datetime.datetime(0))+str(order.data._name)][6] = self.strategy.broker.getposition(order.data).size
            #self.bar_executed = len(self)

        elif order.status in [order.Canceled, order.Margin, order.Rejected]:
            pass
        
        
    def notify_trade(self, trade):
        if trade.justopened:
            pass
            #self.tradetime += 1
            #self.rets.total.open += 1
        elif trade.isclosed:
            self.trade_image.append(trade.pnlcomm)
            self.trade_count += 1
            #self.trade_image.append(trade.pnl)


    def next(self):
        super().next()
        self.pass_count +=1
        try:
            self.strategy.params.pass_window
        except AttributeError:
            self.strategy.params.pass_window = 1
        if self.pass_count < self.strategy.params.pass_window:
            pass
        else:
            if self.frequency == 'day':
                self.tradetime.append(self.datas[0].datetime.date(0))
                #self.position_daily[self.datas[0].datetime.date(0)] = self.strategy.broker.getposition(self.datas[0]).size 
                #self.cash_daily[self.datas[0].datetime.date(0)] = self.strategy.broker.get_cash() 
                self.value_daily[self.datas[0].datetime.date(0)] = self.strategy.broker.get_value() 
                self.index_daily[self.datas[0].datetime.date(0)] = self.data.close[0] 
                self.position_cw[self.datas[0].datetime.date(0)] = (self.strategy.broker.get_value()-self.strategy.broker.get_cash())/self.strategy.broker.get_value()
            else:
                self.tradetime.append(self.datas[0].datetime.datetime(0))
                #self.tradetimebar.append(self.datas[0].datetime.datetime(0))
                #self.position_daily[self.datas[0].datetime.date(0)] = self.strategy.broker.getposition(self.datas[0]).size 
                #self.cash_daily[self.datas[0].datetime.date(0)] = self.strategy.broker.get_cash() 
                self.value_daily[self.datas[0].datetime.date(0)] = self.strategy.broker.get_value() 
                self.index_daily[self.datas[0].datetime.date(0)] = self.data.close[0] 
                self.position_cw[self.datas[0].datetime.date(0)] = (self.strategy.broker.get_value()-self.strategy.broker.get_cash())/self.strategy.broker.get_value()

        
    def stop(self):
        # 年度收益
        cur_year = -1
        value_start = 0.0
        value_cur = 0.0
        value_end = 0.0

        self.annualReturn_list = list()
        self.annualReturn_dict = OrderedDict()
        self.monthReturn_list = list()
        self.monthReturn_dict = OrderedDict()
        self.dailyReturn_list = list()
        self.dailyReturn_dict = OrderedDict()

        for i in range(len(self.data) - self.strategy.params.pass_window, -1, -1):
            dt = self.data.datetime.date(-i)
            value_cur = self.strategy.stats.broker.value[-i]

            if dt.year > cur_year:
                if cur_year >= 0:
                    annualret = (value_end / value_start) - 1.0
                    self.annualReturn_list.append(annualret)
                    self.annualReturn_dict[cur_year] = annualret

                    # changing between real years, use last value as new start
                    value_start = value_end
                else:
                    # No value set whatsoever, use the currently loaded value
                    value_start = value_cur

                cur_year = dt.year

            # No matter what, the last value is always the last loaded value
            value_end = value_cur

        if cur_year not in self.annualReturn_dict:
            # finish calculating pending data
            annualret = (value_end / value_start) - 1.0
            self.annualReturn_list.append(annualret)
            self.annualReturn_dict[cur_year] = annualret
        
        #月度收益
        cur_year = -1
        cur_month = -1

        value_start = 0.0
        value_cur = 0.0
        value_end = 0.0

        for i in range(len(self.data)  - self.strategy.params.pass_window, -1, -1):
            dt = self.data.datetime.date(-i)
            value_cur = self.strategy.stats.broker.value[-i]

            if 100*dt.year+dt.month > cur_month:
                if cur_month >= 0:
                    annualret2 = (value_end / value_start) - 1.0
                    self.monthReturn_list.append(annualret2)
                    self.monthReturn_dict[cur_month] = annualret2

                    # changing between real years, use last value as new start
                    value_start = value_end
                else:
                    # No value set whatsoever, use the currently loaded value
                    value_start = value_cur

                cur_month = 100*dt.year+dt.month

            # No matter what, the last value is always the last loaded value
            value_end = value_cur

        if cur_month not in self.monthReturn_dict:
            # finish calculating pending data
            annualret2 = (value_end / value_start) - 1.0
            self.monthReturn_list.append(annualret2)
            self.monthReturn_dict[cur_month] = annualret2
            
        #日度收益
        cur_year = -1
        cur_month = -1
        cur_day = -1

        value_start = 0.0
        value_cur = 0.0
        value_end = 0.0

        for i in range(len(self.data)  - self.strategy.params.pass_window, -1, -1):
            dt = self.data.datetime.date(-i)
            value_cur = self.strategy.stats.broker.value[-i]

            if 10000*dt.year+100*dt.month+dt.day > cur_day:
                if cur_day >= 0:
                    annualret2 = (value_end / value_start) - 1.0
                    self.dailyReturn_list.append(annualret2)
                    self.dailyReturn_dict[cur_day] = annualret2

                    # changing between real years, use last value as new start
                    value_start = value_end
                else:
                    # No value set whatsoever, use the currently loaded value
                    value_start = value_cur

                cur_day = 10000*dt.year+100*dt.month+dt.day

            # No matter what, the last value is always the last loaded value
            value_end = value_cur

        if cur_day not in self.dailyReturn_dict:
            # finish calculating pending data
            annualret2 = (value_end / value_start) - 1.0
            self.dailyReturn_list.append(annualret2)
            self.dailyReturn_dict[cur_day] = annualret2

        #净值
        for i in self.value_daily.keys():
            self.net_daily[i] = self.value_daily[i]/list(self.value_daily.values())[0]
        
        #回撤
        value_daily_list = list(self.value_daily.values())
        
        for i in range(0,len(value_daily_list)):
            self.drawdownX.append((value_daily_list[i]-max(value_daily_list[0:i+1]))/max(value_daily_list[0:i+1]))
            
        
        #交易天数
        count_temp = []
        for i in self.value_daily.keys():
            count_temp.append(10000*i.year+100*i.month+i.day)
        self.count = len(list(set(count_temp)))

        
        #总收益，年化收益，日均收益，收益回撤比, 夏普， Sortino，最终持仓
        self.return_all = list(self.value_daily.values())[-1]/list(self.value_daily.values())[0]-1
        #self.return_all_log = math.log(list(self.value_daily.values())[-1]/list(self.value_daily.values())[0])
        self.return_year = self.return_all/self.count*252
        self.return_dailyMean = self.return_all/self.count
        try:
            self.return_drawdown = -self.return_year/min(self.drawdownX)
        except ZeroDivisionError:
            self.return_drawdown = 0
        self.max_drawdown = -min(self.drawdownX)
        self.sharp = np.mean(list(self.dailyReturn_dict.values()))/np.std(list(self.dailyReturn_dict.values()), ddof=1)*np.sqrt(252)
        self.sortino = np.mean(list(self.dailyReturn_dict.values()))/np.std([i for i in self.dailyReturn_dict.values() if i > 0], ddof=1)*np.sqrt(252)

        #最终持仓
        self.final_data = {}
        for datas in self.strategy.datas:
            self.final_data[str(self.tradetime[-1])+str(datas._name)] = [self.tradetime[-1], datas._name, self.strategy.getposition(data = datas).size]
   
         
    def get_analysis(self):
        #胜率与盈亏，交易次数
        try:
            self.winrate = len([x for x in self.trade_image if x > 0])/len(self.trade_image)
            self.pn_ratio = -np.mean([x for x in self.trade_image if x > 0])/np.mean([x for x in self.trade_image if x < 0])
            self.trade_count = len(self.trade_image)
        except ZeroDivisionError:
            self.winrate = 0
            self.pn_ratio = 0
            self.trade_count = len(self.trade_image)
        #年收益
        def get_plot_1():
            from matplotlib.pyplot import MultipleLocator
            plt.rcParams['figure.dpi'] = 150
            plt.rcParams['font.sans-serif'] = ['SimHei']
            colour = ['r' if q>0 else 'g' for q in self.annualReturn_dict.values()]
            fig, ax1 = plt.subplots()
            ax1.set_title('年收益率')
            ax1.xaxis.set_major_locator(MultipleLocator(1))
            ax1.bar([str(x) for x in list(self.annualReturn_dict.keys())], list(self.annualReturn_dict.values()), width = 0.5, color = colour)
            plt.xticks(rotation=30)
            ax1.set_ylabel('增长率')
            ax2 = ax1.twinx()
            ax2.xaxis.set_major_locator(MultipleLocator(1))
            ax2.plot([str(x) for x in list(self.annualReturn_dict.keys())], list(1+np.cumsum(list(self.annualReturn_dict.values()))),color = 'y', linestyle = '-.') 
            ax2.set_ylabel('净值')
            fig.savefig('D:/work/y.png')
        #月收益
        def get_plot_2():
            from matplotlib.pyplot import MultipleLocator
            plt.rcParams['figure.dpi'] = 150
            plt.rcParams['font.sans-serif'] = ['SimHei']
            colour = ['r' if q>0 else 'g' for q in self.monthReturn_dict.values()]
            fig, ax1 = plt.subplots()
            ax1.set_title('月度收益率')
            ax1.xaxis.set_major_locator(MultipleLocator(10))
            ax1.bar([str(x) for x in list(self.monthReturn_dict.keys())], list(self.monthReturn_dict.values()), width = 0.8, color = colour)
            plt.xticks(rotation=30)
            ax1.set_ylabel('增长率')
            ax2 = ax1.twinx()
            ax2.xaxis.set_major_locator(MultipleLocator(10))
            ax2.plot([str(x) for x in list(self.monthReturn_dict.keys())], list(1+np.cumsum(list(self.monthReturn_dict.values()))),color = 'y', linestyle = '-.')         
            ax2.set_ylabel('净值')
            fig.savefig('D:/work/m.png')
        #仓位
        def get_plot_3():
            fig, ax = plt.subplots()
            ax.stackplot(list(self.position_cw.keys()),list(self.position_cw.values()))
            ax.set_title('仓位')
            plt.xticks(rotation=30)
            fig.savefig('D:/work/c.png')
        def get_plot_4():
            fig, ax = plt.subplots()
            ax.plot(list(self.value_daily.keys()),list(self.value_daily.values()))
            ax.grid()
        def get_plot_5():
            from matplotlib.ticker import FuncFormatter
            plt.rcParams['font.sans-serif'] = ['SimHei']
            plt.rcParams['axes.unicode_minus'] = False
            fig, ax = plt.subplots(2,1,gridspec_kw = {'height_ratios': [3, 1]})
            ax[0].plot(list(self.net_daily.keys()),list(self.net_daily.values()))
            ax[0].set_title('净值')
            ax[0].grid()
            plt.xticks(rotation=30)
            ax[1].stackplot(list(self.net_daily.keys()), self.drawdownX)
            ax[1].set_title('回撤')
            ax[1].grid()
            ax[1].yaxis.set_major_formatter(FuncFormatter(lambda x,y:str(100*x)+'%'))
            plt.xticks(rotation=30)
            fig.tight_layout()
            fig.savefig('D:/work/net.png')
        def get_plot_6():
            plt.rcParams['font.sans-serif'] = ['SimHei']
            fig, ax = plt.subplots()
            ax.plot(list(self.net_daily.keys()),list(self.net_daily.values()))
            f = list(self.index_daily.values())[0]
            ax.plot(list(self.net_daily.keys()),[i/f for i in list(self.index_daily.values())])
            plt.xticks(rotation=30)
            ax.set_title('策略与指数净值对比')
            ax.grid()
            fig.savefig('D:/work/compare.png')
      
        def get_excel():
            
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image
            #book = load_workbook('D:/work/result.xlsx')
            writer = pd.ExcelWriter('D:/work/result.xlsx', engine='openpyxl')
            #writer.book = book
            result_df = pd.DataFrame([self.count,self.return_all,self.return_year,self.return_dailyMean,self.max_drawdown,self.sharp,self.sortino,self.return_drawdown,self.winrate,self.pn_ratio,self.trade_count])
            result_df.index = ['交易天数', '总收益', '年化收益', '日均收益', '最大回撤', '年化夏普', '年化sortino', '收益回撤比', '胜率', '盈亏', '交易次数']
            result_df.columns = ['data']
            result_df.to_excel(writer, sheet_name='摘要')
            net_df = pd.Series(self.net_daily)
            net_df.name = 'net'
            net_df.to_excel(writer, sheet_name='净值')
            order_df = pd.DataFrame(self.order_image).T
            order_df.columns = ['o_datetime','o_code','o_price','o_volume','o_fee','o_margin','share']
            order_df.to_excel(writer, sheet_name='订单流')
            finaldata_df = pd.DataFrame(self.final_data).T
            finaldata_df.columns = ['date', 'name', 'share']
            finaldata_df.to_excel(writer, sheet_name='最终持仓')
            portfolio_df = order_df.loc[:,['o_datetime','o_code','share']]
            portfolio_df.to_excel(writer, sheet_name='组合列表')
            annualReturn_df = pd.Series(self.annualReturn_dict)
            annualReturn_df.name = 'return'
            annualReturn_df.to_excel(writer, sheet_name='年收益')
            m = {}
            for i in range(int(list(self.monthReturn_dict.keys())[0]/100),int(list(self.monthReturn_dict.keys())[-1]/100)+1):
                m[i] = [np.nan for _ in range(12)]
            monthReturn_df = pd.DataFrame(m).T
            monthReturn_df.columns = [1,2,3,4,5,6,7,8,9,10,11,12]
            for i in list(self.monthReturn_dict.keys()):
                monthReturn_df.at[int(i/100),i%100] = self.monthReturn_dict[i]*100
            mean_df = monthReturn_df.mean(axis = 1)
            std_df = monthReturn_df.std(axis = 1)
            monthReturn_df['mean'] = mean_df
            monthReturn_df['std'] = std_df
            monthReturn_df.to_excel(writer, sheet_name='月收益')
            
            writer.save()
            writer.close()
            
            book = load_workbook('D:/work/result.xlsx')
            img1 = Image('D:/work/net.png')
            img1.width, img1.height = (600, 400)
            img2 = Image('D:/work/compare.png')
            img2.width, img2.height = (600, 400)  
            img3 = Image('D:/work/y.png')
            img3.width, img3.height = (600, 400)  
            img4 = Image('D:/work/m.png')
            img4.width, img4.height = (600, 400)  
            img5 = Image('D:/work/c.png')
            img5.width, img5.height = (600, 400)  
            sheet = book['摘要']
            sheet.add_image(img1, 'D1')
            book.save('D:/work/result.xlsx')
            sheet.add_image(img2, 'M1')
            book.save('D:/work/result.xlsx')
            sheet.add_image(img3, 'D26')
            book.save('D:/work/result.xlsx')
            sheet.add_image(img4, 'M26')
            book.save('D:/work/result.xlsx')
            sheet.add_image(img5, 'V1')
            book.save('D:/work/result.xlsx')
        
         
        get_plot_1()
        get_plot_2()
        get_plot_5()
        get_plot_6()
        get_plot_3()
        #print(self.sharp)
        
        result = pd.DataFrame([self.count,self.return_all,self.return_year,self.return_dailyMean,self.max_drawdown,self.sharp,self.sortino,self.return_drawdown,self.winrate,self.pn_ratio,self.trade_count])
        result.index = ['交易天数', '总收益', '年化收益', '日均收益', '最大回撤', '年化夏普', '年化sortino', '收益回撤比', '胜率', '盈亏', '交易次数']
        get_excel()  
          
        return result.T
                
                
                
                
                
                
                
        